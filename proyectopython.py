import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from colorama import Fore, init, Style

# Inicializar colorama para colores en la terminal
init(autoreset=True)

# Ruta del archivo Excel
EXCEL_FILE = 'COBRO VENDEDOR 24 DE JUNIO 24 beta.xlsm'

# Nombres de las hojas a cargar
hojas = ['CURICO', 'TALCA', 'LINARES']
datos = {}

# Cargar cada hoja comenzando desde la fila 3 (índice 2) para que 'Cliente' esté en la fila correcta
for hoja in hojas:
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=hoja, header=3, engine='openpyxl')
        print(Fore.GREEN + f"Hoja '{hoja}' cargada exitosamente.")
        # Verificar si la columna 'Cliente' existe
        if 'Cliente' in df.columns:
            datos[hoja] = df
            print(Fore.GREEN + f"Datos de la hoja '{hoja}':")
            print(df.head())  # Mostrar las primeras filas para ver los datos cargados
        else:
            print(Fore.RED + f"Advertencia: La columna 'Cliente' no se encontró en la hoja {hoja}.")
    except Exception as e:
        print(Fore.RED + f"Error al cargar la hoja {hoja}: {e}")

print("Hojas cargadas en el diccionario 'datos':", datos.keys())

def limpiar_pantalla():
    os.system('cls' if os.name == 'nt' else 'clear')

def guardar_clientes_bloqueados():
    # Crear DataFrame para clientes bloqueados
    bloqueados = []
    for comuna, df in datos.items():
        if 'Estado' in df.columns:
            bloqueados_comuna = df[df['Estado'] == 'bloqueado'].copy()
            bloqueados_comuna['Comuna'] = comuna  # Añadir la columna Comuna
            bloqueados.extend(bloqueados_comuna[['Comuna', 'Cliente', 'Nº Factura', 'Vencimiento']].values.tolist())

    if bloqueados:
        df_bloqueados = pd.DataFrame(bloqueados, columns=['Comuna', 'Cliente', 'Factura', 'Vencimiento'])
        df_bloqueados.to_excel('Clientes_y_Facturas.xlsx', sheet_name='Clientes Bloqueados', index=False)
        print(Fore.GREEN + "Clientes bloqueados guardados en 'Clientes_y_Facturas.xlsx'.")
    else:
        print(Fore.YELLOW + "No hay clientes bloqueados para guardar.")

def guardar_cambios():
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        for hoja, df in datos.items():
            df.to_excel(writer, sheet_name=hoja, index=False)
        # Guardar los cambios en la hoja Bloqueados
        guardar_clientes_bloqueados()

def crear_excel_formateado():
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes y Facturas"

    # Estilos
    bold_font = Font(bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    # Agregar los datos de todas las hojas
    for hoja, df in datos.items():
        ws.append([f"Datos de {hoja}"])
        ws.append(["Cliente", "Factura", "Vencimiento"])

        # Aplicar formato a los encabezados
        for col in range(1, 4):
            cell = ws.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Verificar si la columna 'Cliente' existe antes de procesar
        if 'Cliente' in df.columns:
            df = df.dropna(subset=['Cliente'])
            df = df[df['Cliente'].str.strip() != '']  # Eliminar filas con 'Cliente' vacío

            # Obtener la lista de clientes, facturas y vencimientos
            clientes = df['Cliente'].tolist()
            facturas = df['Nº Factura'].tolist()
            vencimientos = df['Vencimiento'].tolist()

            # Cambiar el formato de vencimiento
            vencimientos = []
            for v in df['Vencimiento']:
                if pd.notnull(v):
                    if isinstance(v, str):
                        try:
                            v = pd.to_datetime(v, errors='coerce')  # Convertir cadenas a datetime
                        except ValueError:
                            v = pd.NaT
                    if pd.notna(v):
                        vencimientos.append(v.strftime("%Y-%m-%d"))
                    else:
                        vencimientos.append('nan')
                else:
                    vencimientos.append('nan')

            for cliente, factura, vencimiento in zip(clientes, facturas, vencimientos):
                ws.append([cliente, factura, vencimiento])

            # Aplicar borde a cada celda
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.border = border

            ws.append([])  # Añadir una línea en blanco
        else:
            print(Fore.RED + f"La columna 'Cliente' no existe en la hoja {hoja}, se omitirá.")

    # Agregar hoja para clientes bloqueados
    ws_bloqueados = wb.create_sheet(title="Clientes Bloqueados")
    ws_bloqueados.append(["Comuna", "Cliente", "Factura", "Vencimiento"])

    # Aplicar formato a los encabezados de clientes bloqueados
    for col in range(1, 5):
        cell = ws_bloqueados.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Obtener datos de clientes bloqueados
    for comuna, df in datos.items():
        if 'Estado' in df.columns:
            bloqueados = df[df['Estado'] == 'bloqueado'].copy()
            bloqueados['Comuna'] = comuna  # Añadir la columna Comuna
            for _, row in bloqueados.iterrows():
                cliente = row['Cliente']
                factura = row['Nº Factura']
                vencimiento = row['Vencimiento']
                if pd.notnull(vencimiento):
                    if isinstance(vencimiento, str):
                        try:
                            vencimiento = pd.to_datetime(vencimiento, errors='coerce')
                        except ValueError:
                            vencimiento = pd.NaT
                    if pd.notna(vencimiento):
                        vencimiento = vencimiento.strftime("%Y-%m-%d")
                    else:
                        vencimiento = 'nan'
                else:
                    vencimiento = 'nan'
                ws_bloqueados.append([comuna, cliente, factura, vencimiento])

    # Aplicar borde a cada celda en la hoja de clientes bloqueados
    for row in ws_bloqueados.iter_rows(min_row=2, max_row=ws_bloqueados.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border

    # Aplicar formato a todas las celdas
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = Font(color="00FF00")
                cell.alignment = cell.alignment.copy(horizontal="left")

        # Ajustar el ancho de las columnas automáticamente
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Guardar el archivo
    wb.save('Clientes_y_Facturas.xlsx')
    print(Fore.GREEN + "Archivo Excel formateado creado con éxito.")

def mostrar_menu():
    limpiar_pantalla()
    print(Fore.BLUE + Style.BRIGHT + "===== MENÚ PRINCIPAL =====")
    print(Fore.CYAN + "1. Mostrar clientes")
    print(Fore.CYAN + "2. Bloquear cliente")
    print(Fore.CYAN + "3. Desbloquear cliente")
    print(Fore.CYAN + "4. Eliminar cliente")
    print(Fore.CYAN + "5. Crear archivo Excel formateado")
    print(Fore.CYAN + "6. Salir")

def mostrar_submenu():
    limpiar_pantalla()
    print(Fore.BLUE + Style.BRIGHT + "===== SELECCIÓN DE COMUNA =====")
    print(Fore.CYAN + "1. CURICO")
    print(Fore.CYAN + "2. TALCA")
    print(Fore.CYAN + "3. LINARES")
    print(Fore.CYAN + "4. Salir")

    seleccion = input(Fore.GREEN + "Seleccione una comuna: ")
    return seleccion

def mostrar_clientes(comuna):
    if comuna in datos:
        df = datos[comuna]
        if 'Cliente' in df.columns:
            print(Fore.GREEN + f"Clientes en la comuna '{comuna}':")
            print(df[['Cliente', 'Nº Factura', 'Vencimiento']].to_string(index=False))
        else:
            print(Fore.RED + f"La columna 'Cliente' no se encuentra en la hoja {comuna}.")
    else:
        print(Fore.RED + f"No hay datos para la comuna '{comuna}'.")

def bloquear_cliente(nombre, comuna):
    if comuna in datos:
        df = datos[comuna]
        if 'Cliente' in df.columns:
            df.loc[df['Cliente'].str.strip() == nombre, 'Estado'] = 'bloqueado'
            print(Fore.GREEN + f"Cliente '{nombre}' bloqueado en la comuna '{comuna}'.")
        else:
            print(Fore.RED + f"La columna 'Cliente' no se encuentra en la hoja {comuna}.")
    else:
        print(Fore.RED + f"No hay datos para la comuna '{comuna}'.")

def desbloquear_cliente(nombre, comuna):
    if comuna in datos:
        df = datos[comuna]
        if 'Cliente' in df.columns:
            df.loc[df['Cliente'].str.strip() == nombre, 'Estado'] = 'activo'
            print(Fore.GREEN + f"Cliente '{nombre}' desbloqueado en la comuna '{comuna}'.")
        else:
            print(Fore.RED + f"La columna 'Cliente' no se encuentra en la hoja {comuna}.")
    else:
        print(Fore.RED + f"No hay datos para la comuna '{comuna}'.")

def eliminar_cliente(nombre, comuna):
    if comuna in datos:
        df = datos[comuna]
        if 'Cliente' in df.columns:
            df = df[df['Cliente'].str.strip() != nombre]
            datos[comuna] = df
            print(Fore.GREEN + f"Cliente '{nombre}' eliminado de la comuna '{comuna}'.")
        else:
            print(Fore.RED + f"La columna 'Cliente' no se encuentra en la hoja {comuna}.")
    else:
        print(Fore.RED + f"No hay datos para la comuna '{comuna}'.")

def main():
    while True:
        mostrar_menu()
        opcion = input(Fore.GREEN + "\nSeleccione una opción: ")
        if opcion == '1':
            comuna = mostrar_submenu()
            if comuna == '4':
                continue
            nombre_hoja = hojas[int(comuna) - 1] if comuna in ['1', '2', '3'] else None
            if nombre_hoja:
                mostrar_clientes(nombre_hoja)
            else:
                print(Fore.RED + "Opción no válida. Intente de nuevo.")
        elif opcion == '2':
            comuna = mostrar_submenu()
            if comuna == '4':
                continue
            nombre = input(Fore.GREEN + "Ingrese el nombre del cliente a bloquear: ")
            nombre_hoja = hojas[int(comuna) - 1] if comuna in ['1', '2', '3'] else None
            if nombre_hoja:
                bloquear_cliente(nombre, nombre_hoja)
            else:
                print(Fore.RED + "Opción no válida. Intente de nuevo.")
        elif opcion == '3':
            comuna = mostrar_submenu()
            if comuna == '4':
                continue
            nombre = input(Fore.GREEN + "Ingrese el nombre del cliente a desbloquear: ")
            nombre_hoja = hojas[int(comuna) - 1] if comuna in ['1', '2', '3'] else None
            if nombre_hoja:
                desbloquear_cliente(nombre, nombre_hoja)
            else:
                print(Fore.RED + "Opción no válida. Intente de nuevo.")
        elif opcion == '4':
            comuna = mostrar_submenu()
            if comuna == '4':
                continue
            nombre = input(Fore.GREEN + "Ingrese el nombre del cliente a eliminar: ")
            nombre_hoja = hojas[int(comuna) - 1] if comuna in ['1', '2', '3'] else None
            if nombre_hoja:
                eliminar_cliente(nombre, nombre_hoja)
            else:
                print(Fore.RED + "Opción no válida. Intente de nuevo.")
        elif opcion == '5':
            crear_excel_formateado()
        elif opcion == '6':
            guardar_cambios()
            print(Fore.YELLOW + "Saliendo...")
            break
        else:
            print(Fore.RED + "Opción no válida. Intente de nuevo.")

if __name__ == '__main__':
    main()
